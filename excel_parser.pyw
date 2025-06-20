# motdfile = "motd/<gamemode>txt"
# Need to remake these...
from itertools import chain
import openpyxl
from path import Path


DOCUMENT_FILE = Path("F:\\Projects\\gamemodes\\GunGame Data.xlsx")
SERVER_DIRECTORY = Path("E:\\Servers\\cstrike\\cstrike")
GAMEMODE_DIRECTORY = SERVER_DIRECTORY / "cfg" / "gamemodes"
MAPS_DIRECTORY = SERVER_DIRECTORY / "cfg" / "maps"
GUNGAME_PLUGIN_PATH = Path(
    "F:\\Plugins\\gungame\\addons\\source-python\\plugins\\gungame\\plugins"
)
PLUGIN_LISTS = {
    "gg": {
        "included gg plugins": [
            str(directory.stem)[3:]
            for directory in (GUNGAME_PLUGIN_PATH / 'included').dirs("[!_]*")
        ],
        "custom gg plugins": [
            str(directory.stem)[3:]
            for directory in (GUNGAME_PLUGIN_PATH / 'custom').dirs("[!_]*")
        ]
    },
    "sp": {
        "plugins": [
            str(directory.stem)
            for directory in SERVER_DIRECTORY.joinpath(
                "addons",
                "source-python",
                "plugins",
            ).dirs()
        ],
    },
}
CONVERSIONS = {
    "Free for All": "FFA"
}


class GameModeParser:
    def __init__(self, map_data, plugin_data, config_data):
        self.map_data = map_data
        self.plugin_data = plugin_data
        self.config_data = config_data
        self.map_specific_plugins = [
            key for key, value in chain.from_iterable(
                outer_dict.get("Map Specific", {}).items()
                for outer_dict in self.plugin_data.values()
            ) if value
        ]
        self.map_data_keys = set(filter(None, self.map_data))
        self.validate_data()

    def validate_data(self):
        assert self.map_data_keys.issubset(self.plugin_data["sp"])
        assert self.map_data_keys.issubset(self.plugin_data["gg"])
        assert self.map_data_keys.issubset(self.config_data)

    def parse_data(self):
        for gamemode in sorted(self.map_data_keys):
            map_name = self.map_data[gamemode]
            map_file = MAPS_DIRECTORY / map_name + ".cfg"
            gamemode_key = gamemode
            for key, value in CONVERSIONS.items():
                gamemode_key = gamemode_key.replace(key, value)
            gamemode_key = "_".join(gamemode_key.split()).lower()
            contents, map_specific = self.get_data_for_gamemode(gamemode)
            contents += f'motdfile "motd/{gamemode_key}.txt"\n'
            map_file_contents = f"exec gamemodes/{gamemode_key}\n"
            map_file_contents += "\n".join(map_specific) + "\n"
            with map_file.open("w") as open_file:
                open_file.write(map_file_contents)
            gamemode_file = GAMEMODE_DIRECTORY / gamemode_key + ".cfg"
            with gamemode_file.open("w") as open_file:
                open_file.write(contents)

    def get_data_for_gamemode(self, gamemode):
        contents = ""
        map_specific = []
        for command, values in PLUGIN_LISTS.items():
            for section, plugin_list in values.items():
                data = self.get_data_for_section(
                    plugin_data=self.plugin_data[command][gamemode],
                    plugin_list=plugin_list,
                )
                map_specific.extend([
                    f"{command} plugin load {plugin_name}"
                    for plugin_name in data["map_specific"]
                ])
                messages = {
                    "unload": "",
                    "load": "",
                }
                for key in messages:
                    messages[key] += f"// {key.title()} {section}\n"
                    messages[key] += "\n".join([
                        f"{command} plugin {key} {plugin_name}"
                        for plugin_name in data[key]
                    ])
                    messages[key] += "\n"
                for value in messages.values():
                    contents += value + "\n"
        messages = {
            "gg": "// GunGame variables\n",
            "bot": "// Bot variables\n",
            "other": "// Server variables\n",
        }
        for convar, value in self.config_data[gamemode].items():
            try:
                value = int(value)
            except ValueError:
                value = f'"{value}"'
            key = convar.split("_", 1)[0]
            messages[
                key if key in messages else "other"
            ] += f"{convar} {value}\n"
        contents += "\n".join(messages.values())
        return contents, map_specific

    def get_data_for_section(self, plugin_data, plugin_list):
        return_data = {
            "load": [],
            "unload": [],
            "map_specific": [],
        }
        for plugin_name in plugin_list:
            if plugin_name not in plugin_data:
                continue
            if plugin_name in self.map_specific_plugins:
                if plugin_data[plugin_name]:
                    return_data["map_specific"].append(plugin_name)
            elif plugin_data[plugin_name]:
                return_data["load"].append(plugin_name)
            else:
                return_data["unload"].append(plugin_name)
        return return_data


def get_data_from_sheet(sheet):
    headers = [cell.value for cell in sheet[1][1:]]
    data = {}
    for row in sheet.iter_rows(min_row=2, values_only=True):
        data[row[0]] = dict(zip(headers, row[1:]))

    return data


def get_map_data(sheet):
    return {
        row[0]: row[1]
        for row in sheet.iter_rows(min_row=2, values_only=True)
    }


def run():
    wb = openpyxl.load_workbook(DOCUMENT_FILE)
    map_data = get_map_data(wb["Maps"])
    plugin_data = {
        "sp": get_data_from_sheet(wb["Plugins"]),
        "gg": get_data_from_sheet(wb["GunGame Plugins"]),
    }
    config_data = get_data_from_sheet(wb["Configurations"])
    parser = GameModeParser(
        map_data=map_data,
        plugin_data=plugin_data,
        config_data=config_data,
    )
    parser.parse_data()


if __name__ == "__main__":
    run()
